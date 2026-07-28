from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple


SUPPORTED_TABLE_RULE_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xls", ".docx"}

_HEADER_ALIASES: Dict[str, Tuple[str, ...]] = {
    "target_field": ("targetfield", "target", "targetname", "目标字段", "目标字段名", "目标", "写入字段", "目的字段"),
    "source_fields": ("sourcefield", "sourcefields", "source", "源字段", "源字段名", "来源字段", "输入字段", "原字段"),
    "formula": ("formula", "rule", "expression", "mappingrule", "转换公式", "公式", "规则", "表达式", "映射关系", "转换关系"),
    "conversion_mode": ("conversionmode", "mode", "type", "转换类型", "映射类型", "规则类型"),
    "description": ("description", "desc", "remark", "remarks", "comment", "说明", "备注", "描述", "依据"),
    "concept_name": ("concept", "conceptname", "业务概念", "概念", "字段含义"),
    "source_protocol_type": ("sourceprotocol", "sourceprotocoltype", "源协议", "源协议类型"),
    "source_message_code": ("sourcemessage", "sourcemessagecode", "源消息", "源消息类型", "源消息码"),
    "target_protocol_type": ("targetprotocol", "targetprotocoltype", "目标协议", "目标协议类型"),
    "target_message_code": ("targetmessage", "targetmessagecode", "目标消息", "目标消息类型", "目标消息码"),
}
_MAPPING_PATTERN = re.compile(r"-?\d+(?:\.\d+)?\s*(?:=|->|→)\s*[^,;\n]+")
_ROW_TOKEN_PATTERN = re.compile(r"[A-Za-z_\u4e00-\u9fff][A-Za-z0-9_\u4e00-\u9fff./-]*")
_RESERVED_FORMULA_TOKENS = {
    "if",
    "else",
    "elif",
    "and",
    "or",
    "not",
    "true",
    "false",
    "none",
    "int",
    "float",
    "round",
    "abs",
    "min",
    "max",
    "sum",
    "len",
    "range",
    "enumerate",
    "list",
    "dict",
    "signed",
    "unsigned",
    "scale",
    "clip",
    "result",
    "value",
    "raw",
    "bits",
}


def extract_table_rules_from_files(
    file_paths: Sequence[Any],
    *,
    default_source_protocol_type: Optional[str] = None,
    default_source_message_code: Optional[str] = None,
    default_target_protocol_type: Optional[str] = None,
    default_target_message_code: Optional[str] = None,
) -> Dict[str, Any]:
    """Extract mapping rules from docx/xlsx/xls/csv tables."""
    resolved_files = _normalize_file_paths(file_paths)
    if not resolved_files:
        raise ValueError("table_rule_files不能为空")

    extracted_rules: List[Dict[str, Any]] = []
    file_summaries: List[Dict[str, Any]] = []
    warnings: List[str] = []
    used_relation_ids: set[str] = set()

    for file_path in resolved_files:
        relation_id = _allocate_relation_id(file_path, used_relation_ids)
        tables = _load_file_tables(file_path)
        file_rule_count = 0
        table_summaries: List[Dict[str, Any]] = []
        for table_index, table in enumerate(tables, start=1):
            rules, table_warning = _extract_rules_from_rows(
                table.get("rows") or [],
                file_path=file_path,
                relation_id=relation_id,
                table_label=str(table.get("table_label") or f"table_{table_index}"),
                default_source_protocol_type=default_source_protocol_type,
                default_source_message_code=default_source_message_code,
                default_target_protocol_type=default_target_protocol_type,
                default_target_message_code=default_target_message_code,
            )
            if table_warning:
                warnings.append(table_warning)
            if rules:
                extracted_rules.extend(rules)
                file_rule_count += len(rules)
            table_summaries.append(
                {
                    "table_label": str(table.get("table_label") or f"table_{table_index}"),
                    "row_count": len(table.get("rows") or []),
                    "rule_count": len(rules),
                }
            )

        file_summaries.append(
            {
                "file_path": str(file_path),
                "file_name": file_path.name,
                "relation_id": relation_id,
                "table_count": len(tables),
                "rule_count": file_rule_count,
                "tables": table_summaries,
            }
        )

    if not extracted_rules:
        warning_text = f"，警告: {'; '.join(warnings)}" if warnings else ""
        raise ValueError(f"未从表格文件中提取到有效转换规则{warning_text}")

    return {
        "rules": _dedupe_rules(extracted_rules),
        "file_summaries": file_summaries,
        "warnings": warnings,
    }


def _normalize_file_paths(file_paths: Sequence[Any]) -> List[Path]:
    normalized: List[Path] = []
    for item in file_paths:
        text = str(item or "").strip()
        if not text:
            continue
        path = Path(text).resolve()
        if not path.exists() or not path.is_file():
            raise ValueError(f"表格文件不存在: {path}")
        suffix = path.suffix.lower()
        if suffix not in SUPPORTED_TABLE_RULE_EXTENSIONS:
            raise ValueError(f"暂不支持的表格文件类型: {path.name}")
        normalized.append(path)
    return normalized


def _allocate_relation_id(file_path: Path, used_relation_ids: set[str]) -> str:
    seed = re.sub(r"[^A-Za-z0-9]+", "_", file_path.stem).strip("_") or "table_rules"
    relation_id = seed
    suffix = 2
    while relation_id in used_relation_ids:
        relation_id = f"{seed}_{suffix}"
        suffix += 1
    used_relation_ids.add(relation_id)
    return relation_id


def _load_file_tables(file_path: Path) -> List[Dict[str, Any]]:
    suffix = file_path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return [{"table_label": file_path.name, "rows": _load_csv_rows(file_path, suffix)}]
    if suffix in {".xlsx", ".xls"}:
        return _load_excel_tables(file_path, suffix)
    if suffix == ".docx":
        return _load_docx_tables(file_path)
    raise ValueError(f"暂不支持的表格文件类型: {file_path.name}")


def _load_csv_rows(file_path: Path, suffix: str) -> List[List[str]]:
    encodings = ("utf-8-sig", "utf-8", "gb18030")
    last_error: Optional[Exception] = None
    for encoding in encodings:
        try:
            sample = file_path.read_text(encoding=encoding, errors="strict")
            break
        except Exception as exc:
            last_error = exc
    else:
        raise ValueError(f"读取 CSV 文件失败: {file_path}") from last_error

    delimiter = "\t" if suffix == ".tsv" else ","
    try:
        dialect = csv.Sniffer().sniff(sample[:2048], delimiters=",\t;|")
        delimiter = getattr(dialect, "delimiter", delimiter) or delimiter
    except csv.Error:
        pass
    reader = csv.reader(sample.splitlines(), delimiter=delimiter)
    return [_clean_row(row) for row in reader if any(str(cell or "").strip() for cell in row)]


def _load_excel_tables(file_path: Path, suffix: str) -> List[Dict[str, Any]]:
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - environment dependent
            raise ValueError("当前环境缺少 openpyxl，无法解析 xlsx 文件") from exc

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        try:
            tables = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = [
                    _clean_row(list(row))
                    for row in sheet.iter_rows(values_only=True)
                    if any(str(cell or "").strip() for cell in row)
                ]
                if rows:
                    tables.append({"table_label": sheet_name, "rows": rows})
            return tables
        finally:
            workbook.close()

    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise ValueError("当前环境缺少 xlrd，无法解析 xls 文件") from exc

    workbook = xlrd.open_workbook(file_path)
    tables = []
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        rows = [
            _clean_row([sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)])
            for row_index in range(sheet.nrows)
        ]
        rows = [row for row in rows if any(cell for cell in row)]
        if rows:
            tables.append({"table_label": sheet.name, "rows": rows})
    return tables


def _load_docx_tables(file_path: Path) -> List[Dict[str, Any]]:
    try:
        from docx import Document
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise ValueError("当前环境缺少 python-docx，无法解析 docx 文件") from exc

    document = Document(file_path)
    tables = []
    for index, table in enumerate(document.tables, start=1):
        rows = []
        for row in table.rows:
            cleaned = _clean_row([cell.text for cell in row.cells])
            if any(cleaned):
                rows.append(cleaned)
        if rows:
            tables.append({"table_label": f"table_{index}", "rows": rows})
    return tables


def _clean_row(values: Iterable[Any]) -> List[str]:
    return [str(value or "").strip() for value in values]


def _extract_rules_from_rows(
    rows: Sequence[Sequence[str]],
    *,
    file_path: Path,
    relation_id: str,
    table_label: str,
    default_source_protocol_type: Optional[str],
    default_source_message_code: Optional[str],
    default_target_protocol_type: Optional[str],
    default_target_message_code: Optional[str],
) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    header_row_index, header_map = _detect_header_map(rows)
    if header_row_index is None or "target_field" not in header_map or (
        "source_fields" not in header_map and "formula" not in header_map
    ):
        return [], f"{file_path.name}/{table_label} 未识别到可用的目标字段列和源字段/公式列"

    extracted_rules: List[Dict[str, Any]] = []
    for row_offset, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
        target_field = _read_cell(row, header_map.get("target_field"))
        if not target_field:
            continue
        source_fields = _split_source_fields(_read_cell(row, header_map.get("source_fields")))
        formula = _read_cell(row, header_map.get("formula"))
        if not formula and len(source_fields) == 1:
            formula = source_fields[0]
        if not source_fields and formula and not _looks_like_mapping_table(formula):
            source_fields = _derive_source_fields_from_formula(formula, target_field)
        if not formula or not source_fields:
            continue

        description = _compose_description(
            _read_cell(row, header_map.get("description")),
            file_name=file_path.name,
            table_label=table_label,
            row_number=row_offset,
        )
        extracted_rules.append(
            {
                "message_bundle_id": relation_id,
                "concept_name": _read_cell(row, header_map.get("concept_name")) or target_field,
                "target_field": target_field,
                "source_fields": source_fields,
                "formula": formula,
                "conversion_mode": _infer_conversion_mode(
                    formula,
                    _read_cell(row, header_map.get("conversion_mode")),
                ),
                "description": description,
                "source_protocol_type": _read_cell(row, header_map.get("source_protocol_type")) or default_source_protocol_type,
                "message_code": _read_cell(row, header_map.get("source_message_code")) or default_source_message_code,
                "target_protocol_type": _read_cell(row, header_map.get("target_protocol_type")) or default_target_protocol_type,
                "target_message_code": _read_cell(row, header_map.get("target_message_code")) or default_target_message_code,
                "source": "table_rule_extractor",
                "status": "candidate",
                "table_source": {
                    "file_path": str(file_path),
                    "table_label": table_label,
                    "row_number": row_offset,
                },
            }
        )
    return _dedupe_rules(extracted_rules), None


def _detect_header_map(rows: Sequence[Sequence[str]]) -> Tuple[Optional[int], Dict[str, int]]:
    best_index: Optional[int] = None
    best_map: Dict[str, int] = {}
    best_score = 0
    for row_index, row in enumerate(rows[: min(len(rows), 3)]):
        current_map: Dict[str, int] = {}
        for cell_index, cell in enumerate(row):
            role = _match_header_role(cell)
            if role and role not in current_map:
                current_map[role] = cell_index
        score = len(current_map)
        if "target_field" in current_map:
            score += 3
        if "source_fields" in current_map:
            score += 2
        if "formula" in current_map:
            score += 2
        if score > best_score:
            best_index = row_index
            best_map = current_map
            best_score = score
    return best_index, best_map


def _match_header_role(text: str) -> Optional[str]:
    normalized = _normalize_header_text(text)
    if not normalized:
        return None
    for role, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            if alias == normalized or alias in normalized:
                return role
    return None


def _normalize_header_text(text: Any) -> str:
    normalized = str(text or "").strip().lower()
    normalized = re.sub(r"[\s_\-./:：()（）\[\]【】]+", "", normalized)
    return normalized


def _read_cell(row: Sequence[str], index: Optional[int]) -> str:
    if index is None or index >= len(row):
        return ""
    return str(row[index] or "").strip()


def _split_source_fields(value: str) -> List[str]:
    text = str(value or "").strip()
    if not text:
        return []
    parts = re.split(r"[\n,，;；、|/]+", text)
    normalized: List[str] = []
    seen = set()
    for part in parts:
        candidate = str(part or "").strip()
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        normalized.append(candidate)
    return normalized


def _derive_source_fields_from_formula(formula: str, target_field: str) -> List[str]:
    target_tokens = {str(target_field or "").strip(), _to_formula_token(target_field)}
    inferred: List[str] = []
    seen = set()
    for token in _ROW_TOKEN_PATTERN.findall(str(formula or "")):
        cleaned = str(token or "").strip()
        if not cleaned:
            continue
        if cleaned.lower() in _RESERVED_FORMULA_TOKENS:
            continue
        if cleaned in target_tokens:
            continue
        if cleaned in seen:
            continue
        if re.fullmatch(r"-?\d+(?:\.\d+)?", cleaned):
            continue
        seen.add(cleaned)
        inferred.append(cleaned)
    return inferred


def _compose_description(raw_description: str, *, file_name: str, table_label: str, row_number: int) -> str:
    source_marker = f"来源文件 {file_name} / {table_label} / 第{row_number}行"
    description = str(raw_description or "").strip()
    if not description:
        return source_marker
    if source_marker in description:
        return description
    return f"{description}；{source_marker}"


def _infer_conversion_mode(formula: str, explicit_mode: str) -> str:
    explicit = str(explicit_mode or "").strip().lower()
    if explicit in {"mapping", "transcoding"}:
        return explicit
    return "mapping" if _looks_like_mapping_table(formula) else "transcoding"


def _looks_like_mapping_table(formula: str) -> bool:
    text = str(formula or "").strip()
    if not text:
        return False
    if "==" in text or " if " in text or text.startswith("if "):
        return False
    if any(operator in text for operator in ("+", "*", "/", "%")):
        return False
    return bool(_MAPPING_PATTERN.search(text))


def _dedupe_rules(rules: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    deduped: List[Dict[str, Any]] = []
    seen = set()
    for rule in rules:
        target_field = str(rule.get("target_field") or "").strip()
        source_fields = tuple(str(item or "").strip() for item in (rule.get("source_fields") or []) if str(item or "").strip())
        formula = str(rule.get("formula") or "").strip()
        relation_id = str(rule.get("message_bundle_id") or "").strip()
        key = (relation_id, target_field, source_fields, formula)
        if not target_field or not source_fields or not formula or key in seen:
            continue
        seen.add(key)
        deduped.append(dict(rule))
    return deduped


def _to_formula_token(value: Any) -> str:
    token = re.sub(r"\W+", "_", str(value or "").strip(), flags=re.UNICODE).strip("_")
    if not token:
        return "field"
    if token[0].isdigit():
        token = f"f_{token}"
    return token
